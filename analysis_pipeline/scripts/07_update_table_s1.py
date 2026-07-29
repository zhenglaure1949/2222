from pathlib import Path
import argparse, json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

p=argparse.ArgumentParser()
p.add_argument('--source-dir',default='supplement')
p.add_argument('--out',default='supplement/TableS1_v9_RIPS_gene_sources_runtime_verified.xlsx')
p.add_argument('--results-root',default='results')
a=p.parse_args(); root=Path(a.results_root); src=Path(a.source_dir)
genes=pd.read_csv(src/'gene_sources.tsv',sep='\t').fillna('')
refs=pd.read_csv(src/'references.tsv',sep='\t').fillna('')
rules=pd.read_csv(src/'curation_rules.tsv',sep='\t').fillna('')
cov127=pd.read_csv(root/'gse127136/GSE127136_module_gene_coverage.csv')
cov286=pd.read_csv(root/'gse286911/GSE286911_module_gene_coverage_by_sample.csv')
cov220=pd.read_csv(root/'gse220100/GSE220100_module_gene_coverage.csv')

def module_key(name): return name.replace(' ','_') if name in ['Inflammatory chemotaxis','Endothelial activation','Tubular injury'] else name
def detected(df,module):
 vals=df.loc[df.module==module_key(module),'detected_genes'].dropna().astype(str); out=set()
 for v in vals: out.update(x for x in v.split(';') if x)
 return out
modules=['Fibrosis','Inflammatory chemotaxis','Complement','Endothelial activation','Tubular injury']; sets={}
for m in modules:
 sets[(m,'GSE127136')]=detected(cov127,m); sets[(m,'GSE286911')]=detected(cov286,m); sets[(m,'GSE220100')]=detected(cov220,m)
for i,r in genes.iterrows():
 m,g=r['Module'],r['Gene']
 genes.at[i,'GSE127136 coverage']='Yes' if g in sets[(m,'GSE127136')] else 'No'
 genes.at[i,'GSE286911 coverage']='Yes' if g in sets[(m,'GSE286911')] else 'No'
 genes.at[i,'GSE220100 current-output coverage']='Yes' if g in sets[(m,'GSE220100')] else 'No'
wb=Workbook(); ws=wb.active; ws.title='Gene_sources'
ws.append(['Table S1. Renal injury program score (RIPS): complete gene list, module assignment, rationale, and source traceability'])
ws.append(['RIPS is a broad downstream renal-injury framework, not an IgA-specific biomarker. Coverage was runtime-verified against the real GSE127136, GSE286911, and GSE220100 inputs.']); ws.append([]); ws.append(genes.columns.tolist())
for row in genes.itertuples(index=False,name=None): ws.append(list(row))
ref=wb.create_sheet('References'); ref.append(['Table S1 source bibliography']); ref.append(['PMID, DOI, and URLs are retained for row-wise traceability.']); ref.append([]); ref.append(refs.columns.tolist())
for row in refs.itertuples(index=False,name=None): ref.append(list(row))
cr=wb.create_sheet('Curation_rules'); cr.append(['RIPS curation and analysis rules']); cr.append(['Rules were prespecified to prevent post-hoc gene selection and score overinterpretation.']); cr.append([]); cr.append(rules.columns.tolist())
for row in rules.itertuples(index=False,name=None): cr.append(list(row))
pc=wb.create_sheet('Platform_coverage'); pc.append(['Module-level runtime-verified gene coverage']); pc.append([]); pc.append([]); pc.append(['Module','Rows in Table S1','Unique genes','GSE127136 detected','GSE286911 detected','GSE220100 detected','Full RIPS eligible?','Restricted RIPS eligible?'])
for m in modules:
 sub=genes[genes.Module==m]
 pc.append([m,len(sub),sub.Gene.nunique(),len(sets[(m,'GSE127136')]),len(sets[(m,'GSE286911')]),len(sets[(m,'GSE220100')]),'Yes for whole-transcriptome datasets','Only inflammatory chemotaxis, complement, and endothelial activation for GSE220100'])
vn=wb.create_sheet('Version_notes'); vn.append(['Table S1 version history']); vn.append([]); vn.append([]); vn.append(['Version','Date','Change summary','Status']); vn.append(['v9','2026-07-27','Runtime-verified coverage; formal pyUCell; patient-level pseudobulk; NanoString multi-normalization audit.','Post-analysis submission candidate'])
for sh in wb.worksheets:
 sh.freeze_panes='A5'; sh.sheet_view.showGridLines=False
 for cell in sh[4]: cell.font=Font(bold=True,color='FFFFFF'); cell.fill=PatternFill('solid',fgColor='1F4E78'); cell.alignment=Alignment(wrap_text=True,vertical='center')
 for row in sh.iter_rows():
  for cell in row: cell.alignment=Alignment(wrap_text=True,vertical='top')
 for i in range(1,sh.max_column+1):
  width=min(55,max(12,max((len(str(sh.cell(r,i).value or '')) for r in range(1,sh.max_row+1)),default=12)+2)); sh.column_dimensions[get_column_letter(i)].width=width
out=Path(a.out); out.parent.mkdir(parents=True,exist_ok=True); wb.save(out)
genes.to_csv(out.with_suffix('.tsv'),sep='\t',index=False)
summary={'output':str(out),'rows':len(genes),'references':len(refs),'rules':len(rules),'coverage':{f'{m}_{ds}':len(sets[(m,ds)]) for m in modules for ds in ['GSE127136','GSE286911','GSE220100']}}
(out.parent/'TableS1_v9_runtime_summary.json').write_text(json.dumps(summary,indent=2),encoding='utf-8'); print(json.dumps(summary,indent=2))
